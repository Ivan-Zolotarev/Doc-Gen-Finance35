"""
Генератор Excel документов
"""

import os
from typing import Dict, Any, Optional, List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Генератор документов в формате Excel (.xlsx)"""
    
    def generate(self, template_path: Optional[str], data: Dict[str, Any], output_path: str) -> str:
        """
        Генерация Excel документа
        
        Args:
            template_path: Путь к шаблону Excel (опционально)
            data: Данные для заполнения
            output_path: Путь для сохранения
            
        Returns:
            Путь к сгенерированному файлу
        """
        if template_path and os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = data.get('sheet_name', 'Лист1')
        
        # Заполняем данные
        self._fill_data(ws, data)
        
        # Применяем форматирование
        self._apply_formatting(ws, data)
        
        # Сохранение
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        wb.save(output_path)
        
        return output_path
    
    def _fill_data(self, ws, data: Dict[str, Any]):
        """
        Заполнение данными листа Excel
        
        Args:
            ws: Рабочий лист
            data: Данные для заполнения
        """
        row = 1
        
        # Заголовок
        if 'title' in data:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = data['title']
            cell.font = Font(size=16, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 2
        
        # Таблица данных
        if 'table_data' in data and data['table_data']:
            table_data = data['table_data']
            
            # Заголовки
            for col_idx, header in enumerate(table_data[0], start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = str(header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
            
            # Данные
            for row_data in table_data[1:]:
                for col_idx, cell_data in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = cell_data
                row += 1
        
        # Дополнительные данные
        if 'additional_data' in data:
            row += 1
            for key, value in data['additional_data'].items():
                ws.cell(row=row, column=1).value = str(key)
                ws.cell(row=row, column=2).value = str(value)
                row += 1
    
    def _apply_formatting(self, ws, data: Dict[str, Any]):
        """
        Применение форматирования к листу
        
        Args:
            ws: Рабочий лист
            data: Данные (может содержать настройки форматирования)
        """
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Границы для таблицы
        if 'table_data' in data and data['table_data']:
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Применяем границы ко всем ячейкам с данными
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value:
                        cell.border = thin_border
    
    def create_report(self, data: Dict[str, Any], output_path: str) -> str:
        """
        Создание финансового отчета
        
        Args:
            data: Данные отчета
            output_path: Путь для сохранения
            
        Returns:
            Путь к созданному файлу
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        
        # Заголовок отчета
        ws['A1'] = data.get('report_title', 'Финансовый отчет')
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Дата
        ws['A2'] = f"Дата формирования: {data.get('date', '')}"
        
        # Таблица с финансовыми данными
        if 'financial_data' in data:
            row = 4
            headers = ['Показатель', 'Значение', 'Единица измерения', 'Примечание']
            
            # Заголовки
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            row += 1
            
            # Данные
            for item in data['financial_data']:
                ws.cell(row=row, column=1).value = item.get('name', '')
                ws.cell(row=row, column=2).value = item.get('value', '')
                ws.cell(row=row, column=3).value = item.get('unit', 'руб.')
                ws.cell(row=row, column=4).value = item.get('note', '')
                row += 1
        
        # Применяем форматирование
        self._apply_formatting(ws, data)
        
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        wb.save(output_path)
        
        return output_path

